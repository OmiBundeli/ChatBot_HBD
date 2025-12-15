from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import json

FILE_NAME = "missing_searches.xlsx"

def log_missing_query(query: str, online_results=None):
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["query", "timestamp", "status", "online_snapshot"])
        wb.save(FILE_NAME)

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    ws.append([
        query,
        datetime.utcnow().isoformat(),
        "not_found_in_db",
        json.dumps(online_results[:3]) if online_results else None
    ])

    wb.save(FILE_NAME)
